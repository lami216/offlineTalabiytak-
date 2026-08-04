from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from PIL import Image as PILImage

from app.services.errors import ValidationError
from app.services.excel_pricing import RESULT_HEADERS


def workbook_bytes(headers=("PRICE", "PCS", "CBM"), header_row=1, existing=False):
    book = Workbook()
    sheet = book.active
    sheet.title = "منتجات"
    for column, heading in enumerate(headers, 1):
        sheet.cell(header_row, column, heading)
    if existing:
        for offset, heading in enumerate(RESULT_HEADERS, len(headers) + 1):
            sheet.cell(header_row, offset, heading)
    positions = {}
    for index, heading in enumerate(headers, 1):
        if "PRICE" in heading:
            positions["PRICE"] = index
        if "PCS" in heading and "T.T" not in heading:
            positions["PCS"] = index
        if "CBM" in heading and "T.T" not in heading:
            positions["CBM"] = index
    sheet.cell(header_row + 1, positions["PRICE"], 2.7)
    sheet.cell(header_row + 1, positions["PCS"], 40)
    sheet.cell(header_row + 1, positions["CBM"], 0.091124)
    output = BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


def results(data):
    book = load_workbook(BytesIO(data), data_only=True)
    sheet = book["منتجات"]
    headings = {cell.value: cell.column for cell in sheet[1] if cell.value}
    values = tuple(sheet.cell(2, headings[name]).value for name in RESULT_HEADERS)
    book.close()
    return values


@pytest.mark.asyncio
async def test_pricing_uses_request_parameters_and_exact_calculation(setup):
    _, app, *_ = setup
    source = workbook_bytes()
    first = await app.state.excel_pricing.transform(source, "63", "60000")
    second = await app.state.excel_pricing.transform(source, "65", "70000")
    assert results(first) == pytest.approx((170.1, 5.103, 136.686, 311.889))
    assert results(first) != results(second)


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("headers", "header_row"),
    [
        (("X", "PRICE", "Y", "CBM", "Z", "PCS"), 1),
        (("PCS/CTN", "Extra", "CBM PER CTN", "UNIT PRICE RMB"), 3),
        (("CBM", "PCS", "PRICE"), 8),
    ],
)
async def test_columns_and_header_rows_can_move(setup, headers, header_row):
    _, app, *_ = setup
    output = await app.state.excel_pricing.transform(
        workbook_bytes(headers, header_row), "10", "100"
    )
    book = load_workbook(BytesIO(output), data_only=True)
    sheet = book.active
    assert [
        sheet.cell(header_row, sheet.max_column - 3 + index).value for index in range(4)
    ] == list(RESULT_HEADERS)
    book.close()


@pytest.mark.asyncio
async def test_forbidden_total_aliases_are_not_selected(setup):
    _, app, *_ = setup
    data = workbook_bytes(("AMOUNT", "T.T PCS", "T.T CBM", "PRICE", "PCS", "CBM"))
    output = await app.state.excel_pricing.transform(data, "10", "100")
    assert results(output)[0] == pytest.approx(27)


@pytest.mark.asyncio
async def test_existing_results_are_updated_without_duplicates(setup):
    _, app, *_ = setup
    output = await app.state.excel_pricing.transform(workbook_bytes(existing=True), "10", "100")
    book = load_workbook(BytesIO(output))
    headings = [cell.value for cell in book.active[1]]
    assert all(headings.count(name) == 1 for name in RESULT_HEADERS)
    book.close()


@pytest.mark.asyncio
async def test_results_follow_all_real_data_and_preserve_merged_original_columns(setup):
    _, app, *_ = setup
    book = Workbook()
    sheet = book.active
    sheet.title = "منتجات"
    sheet.append([None] * 16)
    sheet.append([None] * 16)
    sheet.append(
        ["PRICE", "PCS", "CBM"] + [None] * 9 + ["WAREHOUSE", "PRODUCT", "MATERIAL", "NOTE"]
    )
    sheet.append([2, 4, 1] + [None] * 9 + ["A", "one", "steel", "keep"])
    sheet.append([3, 5, 2] + [None] * 10 + ["two", "wood", "also keep"])
    sheet.merge_cells("M4:M5")
    sheet["P4"].fill = PatternFill("solid", fgColor="FF0000")
    pixels = BytesIO()
    PILImage.new("RGB", (2, 2), "blue").save(pixels, format="PNG")
    pixels.seek(0)
    sheet.add_image(Image(pixels), "O6")
    original = {
        (row, col): sheet.cell(row, col).value for row in range(1, 6) for col in range(13, 17)
    }
    buffer = BytesIO()
    book.save(buffer)
    book.close()

    output = await app.state.excel_pricing.transform(buffer.getvalue(), "10", "100")
    result = load_workbook(BytesIO(output), data_only=True)
    sheet = result.active
    assert tuple(sheet.cell(3, col).value for col in range(17, 21)) == RESULT_HEADERS
    assert sheet["Q4"].value == 20
    assert sheet["T4"].value == pytest.approx(45.6)
    preserved = {
        (row, col): sheet.cell(row, col).value for row in range(1, 6) for col in range(13, 17)
    }
    assert preserved == original
    assert "M4:M5" in sheet.merged_cells
    assert sheet["P4"].fill.fgColor.rgb == "00FF0000"
    assert len(sheet._images) == 1
    result.close()


def missing_values_workbook(rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "منتجات"
    sheet.append(["PRICE", "PCS", "CBM"])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


@pytest.mark.asyncio
async def test_missing_and_invalid_components_are_calculated_independently(setup):
    _, app, *_ = setup
    rows = [
        (None, 4, 1),  # shipping only
        (2, 4, None),  # price components only
        (3, None, 1),
        (4, 0, 1),
        (None, None, None),
        (5, 5, 1),
        ("-", 2, 1),
        ("N/A", 4, 2),
        (6, 2, "not a number"),
        (7, 7, 1),
    ]
    output = await app.state.excel_pricing.transform(missing_values_workbook(rows), "10", "100")
    book = load_workbook(BytesIO(output), data_only=True)
    sheet = book.active
    values = [tuple(sheet.cell(row, col).value for col in range(4, 8)) for row in range(2, 12)]
    assert values[0] == pytest.approx((None, None, 25, 25), nan_ok=True)
    assert values[1] == pytest.approx((20, 0.6, None, 20.6), nan_ok=True)
    assert values[2] == pytest.approx((30, 0.9, None, 30.9), nan_ok=True)
    assert values[3] == pytest.approx((40, 1.2, None, 41.2), nan_ok=True)
    assert values[4] == (None, None, None, None)
    assert values[5] == pytest.approx((50, 1.5, 20, 71.5))
    assert values[6] == pytest.approx((None, None, 50, 50), nan_ok=True)
    assert values[7] == pytest.approx((None, None, 50, 50), nan_ok=True)
    assert values[8] == pytest.approx((60, 1.8, None, 61.8), nan_ok=True)
    assert values[9] == pytest.approx((70, 2.1, 100 / 7, 72.1 + 100 / 7))
    book.close()


@pytest.mark.asyncio
async def test_multiple_sheets_and_incomplete_sheet(setup):
    _, app, *_ = setup
    source = workbook_bytes()
    book = load_workbook(BytesIO(source))
    second = book.create_sheet("ثانية")
    second.append(["CBM", "PRICE", "PCS"])
    second.append([1, 2, 4])
    notes = book.create_sheet("ملاحظات")
    notes["A1"] = "unchanged"
    output = BytesIO()
    book.save(output)
    book.close()
    transformed = await app.state.excel_pricing.transform(output.getvalue(), "10", "100")
    result_book = load_workbook(BytesIO(transformed))
    assert result_book["ملاحظات"]["A1"].value == "unchanged"
    assert result_book["ثانية"]["D1"].value == RESULT_HEADERS[0]
    result_book.close()

    incomplete = Workbook()
    incomplete.active.append(["PRICE", "CBM"])
    buffer = BytesIO()
    incomplete.save(buffer)
    with pytest.raises(ValidationError, match="PCS"):
        await app.state.excel_pricing.transform(buffer.getvalue(), "10", "100")


def test_pricing_routes_auth_csrf_validation_and_download(auth):
    client, _, _, _, token, _ = auth
    client.cookies.clear()
    assert client.get("/pricing", follow_redirects=False).status_code == 303
    assert client.post("/pricing", follow_redirects=False).status_code == 303
    client.post("/login", data={"username": "admin", "password": "strong-password"})
    bad = client.post(
        "/pricing",
        files={"file": ("bad.xlsx", b"bad", "application/octet-stream")},
        data={"csrf_token": "bad", "rmb_rate": "1", "shipping_cost_per_cbm": "1"},
    )
    assert bad.status_code == 400
    token = client.app.state.security.load(
        client.cookies[client.app.state.settings.session_cookie_name]
    )["csrf"]
    response = client.post(
        "/pricing",
        files={"file": ("ملف.xlsx", workbook_bytes(), "application/octet-stream")},
        data={"csrf_token": token, "rmb_rate": "63", "shipping_cost_per_cbm": "60000"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    load_workbook(BytesIO(response.content)).close()


@pytest.mark.asyncio
@pytest.mark.parametrize(("rate", "shipping"), [("", "1"), ("0", "1"), ("-1", "1"), ("1", "NaN")])
async def test_parameters_must_be_positive_finite(setup, rate, shipping):
    _, app, *_ = setup
    with pytest.raises(ValidationError):
        await app.state.excel_pricing.transform(workbook_bytes(), rate, shipping)
