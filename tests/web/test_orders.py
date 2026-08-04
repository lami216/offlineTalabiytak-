from datetime import timedelta
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import httpx
import pytest
from openpyxl import load_workbook
from PIL import Image

from app.database.mongo import REQUIRED_INDEXES
from app.main import _asset_version
from app.models import ImageAsset, Order, OrderItem, Product, now
from app.services.excel_export import ExcelExportService
from app.utils.objectid import new_id


async def product(repository, name):
    item = Product(
        new_id(),
        name,
        name,
        ImageAsset("file", f"/{name}.png", "ignored", None, name, "image/png", 20, 20),
    )
    await repository.create(item)
    return item


@pytest.mark.asyncio
async def test_order_repository_lifecycle_and_indexes(setup):
    _, app, _, _, database = setup
    first = await product(app.state.repositories.products, "أول")
    second = await product(app.state.repositories.products, "ثان")
    order = await app.state.orders.create("  طلبية الساحة  ", [first.id, second.id], [2, 7])
    document = database.raw.orders.find_one()
    assert order.title == "طلبية الساحة"
    assert [item.position for item in order.items] == [1, 2]
    assert document["items"][0]["product_id"] != first.id
    assert not any(isinstance(value, bytes) for value in document.values())
    original_expiry = document["expires_at"]
    await app.state.orders.update(order.id, "معدلة", [second.id], [9])
    updated = await app.state.repositories.orders.get_active(order.id)
    assert updated.expires_at == original_expiry
    assert updated.items[0].product_name == "ثان"
    assert "orders_expires_at_ttl" in REQUIRED_INDEXES["orders"]


@pytest.mark.asyncio
async def test_expired_orders_are_hidden(setup):
    _, app, _, _, database = setup
    item = await product(app.state.repositories.products, "قديم")
    order = await app.state.orders.create("منتهية", [item.id], [1])
    database.raw.orders.update_one(
        {"_id": document_id(database, order.id)},
        {"$set": {"expires_at": now() - timedelta(days=1)}},
    )
    assert await app.state.repositories.orders.get_active(order.id) is None
    assert not await app.state.repositories.orders.list_active()


def document_id(database, value):
    return database.raw.orders.find_one({"title": {"$exists": True}})["_id"]


def test_order_routes_excel_and_protection(auth):
    client, app, _, _, token, database = auth
    image = BytesIO()
    with Image.new("RGB", (20, 10), "red") as picture:
        picture.save(image, "PNG")

    def delivery(_request):
        return httpx.Response(200, content=image.getvalue(), headers={"content-type": "image/png"})

    app.state.excel_export.transport = httpx.MockTransport(delivery)
    repository = app.state.repositories.products
    first = client.portal.call(product, repository, "حليب")
    response = client.post(
        "/orders/new",
        data={"csrf_token": token, "title": "طلبية عربية", "product_id": first.id, "quantity": "4"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    download_url = response.headers["location"] + "/download"
    response = client.get(download_url)
    assert response.status_code == 200
    assert response.content.startswith(b"PK")
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["الطلبية"]
    assert sheet.sheet_view.rightToLeft is False
    assert (
        sheet["A1"].value,
        sheet["A3"].value,
        sheet["B3"].value,
        sheet["C3"].value,
        sheet["B4"].value,
        sheet["C4"].value,
    ) == (
        "طلبية عربية",
        "الصورة",
        "اسم المنتج",
        "الكمية",
        "حليب",
        4,
    )
    assert sheet.column_dimensions["A"].width == pytest.approx(9.5)
    assert sheet.column_dimensions["B"].width == pytest.approx(25)
    assert sheet.column_dimensions["C"].width == pytest.approx(9)
    assert sheet.row_dimensions[1].height == 26
    assert sheet.row_dimensions[3].height == 22
    assert sheet.row_dimensions[4].height == 54
    assert sheet["A1"].font.sz == 16
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.color.rgb == "00000000"
    for header in sheet[3]:
        assert header.font.sz == 13
        assert header.font.bold is True
        assert header.font.color.rgb == "00000000"
    assert len(sheet._images) == 1
    workbook.close()
    assert client.get("/orders/product-search?q=x").status_code == 200
    client.cookies.clear()
    assert client.get(download_url, follow_redirects=False).status_code == 303


def test_order_form_includes_quantity_notification_and_script(auth):
    client, *_ = auth

    response = client.get("/orders/new")

    assert response.status_code == 200
    assert "data-order-notification" in response.text
    assert 'role="alert"' in response.text
    assert 'aria-live="assertive"' in response.text
    assert "/static/orders.js?v=" in response.text
    assert "/static/style.css?v=" in response.text
    assert "/static/app.js?v=" in response.text


def test_asset_version_is_content_based(tmp_path):
    asset = tmp_path / "asset.js"
    asset.write_text("first")
    first = _asset_version(asset)
    asset.write_text("second")

    assert len(first) == 12
    assert _asset_version(asset) != first


def test_search_quantity_client_side_contract():
    script = Path("app/static/orders.js").read_text()
    styles = Path("app/static/style.css").read_text()

    assert 'card.className = "card product-search-card"' in script
    assert 'image.className = "product-search-image"' in script
    assert 'name.className = "product-search-name"' in script
    assert 'addControls.className = "product-add-controls"' in script
    assert 'add.className = "product-add-button"' in script
    assert "card.append(image, name, addControls, error)" in script
    assert 'quantityInput.type = "number"' in script
    assert 'quantityInput.min = "1"' in script
    assert 'quantityInput.max = "1000000"' in script
    assert 'quantityInput.step = "1"' in script
    assert 'quantityInput.dataset.addQuantity = ""' in script
    assert "quantityInput.name" not in script
    assert "quantity.value = String(quantityValue)" in script
    assert 'quantity.value = "1"' not in script
    assert 'quantityInput.value = ""' in script
    assert "selected.append(item)" in script
    assert 'event.key === "Enter"' in script
    assert "event.preventDefault()" in script
    assert 'classList.add("is-invalid")' in script
    assert 'setAttribute("aria-invalid", "true")' in script
    assert 'removeAttribute("aria-invalid")' in script
    for message in (
        "رجاءً ضع الكمية المطلوبة.",
        "أدخل كمية صحيحة أكبر من صفر.",
        "هذا المنتج مضاف بالفعل.",
        "تمت إضافة المنتج.",
    ):
        assert message in script

    assert 'grid-template-areas:"image" "name" "controls" "error"' in styles
    assert ".product-search-name{grid-area:name;display:block" in styles
    assert ".product-add-controls{grid-area:controls;display:flex;direction:rtl" in styles
    assert 'quantityInput.className = "product-quantity-input"' in script
    assert 'title.className = "selected-product-name"' in script
    assert 'row.className = "selected-product-row"' in script
    assert 'quantity.className = "selected-quantity-input"' in script
    assert 'controls.className = "actions selected-product-actions"' in script
    assert "flex:0 0 72px;width:72px;min-width:72px;max-width:72px" in styles
    assert "product-quantity-input{flex:1 1 auto" in styles
    assert "selected-quantity-input{flex:0 0 112px;width:112px;min-width:112px" in styles
    assert "flex-basis:104px;width:104px;min-width:104px" in styles
    assert (
        "direction:rtl;text-align:right;unicode-bidi:plaintext;"
        "writing-mode:horizontal-tb;white-space:normal;overflow-wrap:break-word;"
        "word-break:normal" in styles
    )
    assert "text-align:center;direction:ltr" in styles
    assert (
        ".product-quantity-input.is-invalid{border:2px solid #d92d20;"
        "background:#fff5f5;color:#17202a" in styles
    )
    assert ".field-error{grid-area:error" in styles
    assert ".selected-products-list{display:flex;flex-direction:column-reverse" in styles


def test_existing_and_new_selected_quantities_share_markup_contract():
    template = Path("app/templates/order_form.html").read_text()
    script = Path("app/static/orders.js").read_text()

    assert 'class="selected-product-name"' in template
    assert 'class="selected-product-row"' in template
    assert 'class="selected-quantity-input" type="number" name="quantity"' in template
    assert 'class="actions selected-product-actions"' in template
    assert 'row.className = "selected-product-row"' in script
    assert "row.append(quantity, controls)" in script
    assert "item.append(id, title, row)" in script
    assert 'quantity.className = "selected-quantity-input"' in script
    assert 'quantity.type = "number"' in script
    assert 'quantity.name = "quantity"' in script
    assert "quantityInput.name" not in script
    for value in ("1", "50", "1000", "1000000"):
        assert 1 <= int(value) <= 1000000


def test_reversed_visual_order_move_controls_contract():
    script = Path("app/static/orders.js").read_text()

    assert "const next = card.nextElementSibling" in script
    assert "if (next) selected.insertBefore(next, card)" in script
    assert "const previous = card.previousElementSibling" in script
    assert "if (previous) selected.insertBefore(card, previous)" in script


@pytest.mark.asyncio
async def test_order_dom_sequence_stays_excel_sequence_after_edit(setup):
    _, app, _, _, database = setup
    products = [await product(app.state.repositories.products, name) for name in "ABCD"]
    order = await app.state.orders.create(
        "ترتيب المنتجات", [item.id for item in products[:3]], [1, 2, 3]
    )

    assert [item.product_name for item in order.items] == ["A", "B", "C"]
    assert [item.position for item in order.items] == [1, 2, 3]
    document = database.raw.orders.find_one({"title": "ترتيب المنتجات"})
    assert [item["position"] for item in document["items"]] == [1, 2, 3]

    unchanged = await app.state.orders.update(
        order.id, order.title, [item.id for item in products[:3]], [1, 2, 3]
    )
    assert [item.product_name for item in unchanged.items] == ["A", "B", "C"]

    extended = await app.state.orders.update(
        order.id, order.title, [item.id for item in products], [1, 2, 3, 4]
    )
    assert [(item.product_name, item.position) for item in extended.items] == [
        ("A", 1),
        ("B", 2),
        ("C", 3),
        ("D", 4),
    ]

    picture = BytesIO()
    with Image.new("RGB", (10, 10), "green") as source:
        source.save(picture, "PNG")
    workbook = load_workbook(
        BytesIO(ExcelExportService._workbook(extended, [picture.getvalue()] * 4))
    )
    sheet = workbook["الطلبية"]
    assert [sheet.cell(row, 2).value for row in range(4, 8)] == ["A", "B", "C", "D"]
    workbook.close()


@pytest.mark.parametrize("quantity", ["", "0", "-1", "1.5", "1000001"])
def test_order_route_rejects_invalid_quantities(auth, quantity):
    client, app, _, _, token, _ = auth
    item = client.portal.call(product, app.state.repositories.products, "اختبار الكمية")

    response = client.post(
        "/orders/new",
        data={
            "csrf_token": token,
            "title": "طلبية غير صالحة",
            "product_id": item.id,
            "quantity": quantity,
        },
    )

    assert response.status_code == 400


def test_excel_images_crop_to_fill_and_anchor_to_product_rows():
    sizes = ((200, 100), (100, 200), (80, 80))
    pictures = []
    for size in sizes:
        output = BytesIO()
        with Image.new("RGB", size, "green") as source:
            if size[0] > size[1]:
                source.paste("red", (0, 0, 50, size[1]))
                source.paste("blue", (150, 0, size[0], size[1]))
            elif size[1] > size[0]:
                source.paste("red", (0, 0, size[0], 50))
                source.paste("blue", (0, 150, size[0], size[1]))
            source.save(output, "PNG")
        pictures.append(output.getvalue())
    order = Order(
        "order-id",
        "طلبية الصور",
        [OrderItem(f"product-{index}", f"منتج {index}", index, index) for index in range(1, 4)],
    )

    data = ExcelExportService._workbook(order, pictures)

    workbook = load_workbook(BytesIO(data))
    sheet = workbook["الطلبية"]
    assert sheet.sheet_view.rightToLeft is False
    assert sheet.column_dimensions["A"].width == pytest.approx(9.5)
    assert sheet.column_dimensions["B"].width == pytest.approx(25)
    assert sheet.column_dimensions["C"].width == pytest.approx(9)
    assert len(sheet._images) == len(order.items)
    for row, embedded in enumerate(sheet._images, 4):
        assert (embedded.width, embedded.height) == (72, 72)
        assert embedded.anchor._from.col == 0
        assert embedded.anchor._from.row == row - 1
        assert embedded.anchor._from.colOff == 0
        assert embedded.anchor._from.rowOff == 0
        assert embedded.anchor.ext.cx == 72 * 9525
        assert embedded.anchor.ext.cy == 72 * 9525
        assert sheet.row_dimensions[row].height == 54
        assert sheet.cell(row, 2).value == f"منتج {row - 3}"
        assert sheet.cell(row, 2).font.sz == 14
        assert sheet.cell(row, 2).font.bold is True
        assert sheet.cell(row, 2).font.color.rgb == "00000000"
        assert sheet.cell(row, 2).alignment.horizontal == "center"
        assert sheet.cell(row, 2).alignment.vertical == "center"
        assert sheet.cell(row, 2).alignment.wrap_text is True
        assert sheet.cell(row, 2).alignment.readingOrder == 2
        assert sheet.cell(row, 3).value == row - 3
        assert isinstance(sheet.cell(row, 3).value, int)
        assert sheet.cell(row, 3).font.sz == 13
        assert sheet.cell(row, 3).font.bold is True
        assert sheet.cell(row, 3).font.color.rgb == "00000000"
        assert sheet.cell(row, 3).alignment.horizontal == "center"
        assert sheet.cell(row, 3).alignment.vertical == "center"
    workbook.close()

    with ZipFile(BytesIO(data)) as archive:
        media = sorted(name for name in archive.namelist() if name.startswith("xl/media/"))
        assert len(media) == len(sizes)
        for name in media:
            with Image.open(BytesIO(archive.read(name))) as embedded:
                assert embedded.size == (72, 72)
                # Colored outer bands are removed by a centered cover crop, not squeezed in.
                edge_midpoints = ((0, 36), (71, 36), (36, 0), (36, 71))
                for pixel in (embedded.getpixel(point) for point in edge_midpoints):
                    red, green, blue = pixel
                    assert green > red
                    assert green > blue
