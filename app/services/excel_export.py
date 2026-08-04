import asyncio
import re
from io import BytesIO
from urllib.parse import urlparse

import httpx
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image, ImageOps, UnidentifiedImageError

from app.services.errors import AppError, ValidationError


class ExcelExportService:
    def __init__(self, settings, products, transport=None):
        self.settings, self.products, self.transport = settings, products, transport

    async def build(self, order):
        pictures = []
        for item in sorted(order.items, key=lambda i: i.position):
            product = await self.products.get(item.product_id)
            if not product:
                raise ValidationError(
                    "تحتوي الطلبية على منتج لم يعد متاحًا. عدّل الطلبية وأزل المنتج المفقود."
                )
            pictures.append(await self._download(product.primary_image.file_path))
        try:
            return await asyncio.to_thread(self._workbook, order, pictures)
        except AppError:
            raise
        except Exception as exc:
            raise AppError("تعذر إنشاء ملف Excel.") from exc

    async def _download(self, file_path):
        url = self.settings.imagekit_delivery_url(file_path)
        expected = urlparse(self.settings.imagekit_url_endpoint)
        actual = urlparse(url)
        if (actual.scheme, actual.netloc) != (expected.scheme, expected.netloc):
            raise AppError("فشل تنزيل صورة من ImageKit.")
        limit = self.settings.excel_image_max_mb * 1024 * 1024
        try:
            async with httpx.AsyncClient(
                timeout=self.settings.excel_image_timeout_seconds,
                follow_redirects=False,
                transport=self.transport,
            ) as client:
                async with client.stream("GET", url) as response:
                    if response.status_code != 200 or not response.headers.get(
                        "content-type", ""
                    ).lower().startswith("image/"):
                        raise AppError("فشل تنزيل صورة من ImageKit.")
                    data = bytearray()
                    async for chunk in response.aiter_bytes():
                        data.extend(chunk)
                        if len(data) > limit:
                            raise AppError("حجم صورة المنتج يتجاوز الحد المسموح.")
            return bytes(data)
        except AppError:
            raise
        except httpx.HTTPError as exc:
            raise AppError("فشل تنزيل صورة من ImageKit.") from exc

    @staticmethod
    def _workbook(order, pictures):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "الطلبية"
        sheet.sheet_view.rightToLeft = False
        sheet.merge_cells("A1:C1")
        sheet["A1"] = order.title
        sheet["A1"].font = Font(name="Arial", bold=True, size=16, color="000000")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        sheet.row_dimensions[1].height = 26
        headers = ("الصورة", "اسم المنتج", "الكمية")
        border = Border(*([Side(style="thin", color="888888")] * 4))
        for col, value in enumerate(headers, 1):
            cell = sheet.cell(3, col, value)
            cell.font = Font(name="Arial", bold=True, size=13, color="000000")
            cell.fill = PatternFill("solid", fgColor="DDEFE9")
            cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
            cell.border = border
        sheet.row_dimensions[3].height = 22
        keepalive = []
        try:
            for row, (item, raw) in enumerate(
                zip(sorted(order.items, key=lambda i: i.position), pictures, strict=True), 4
            ):
                converted = BytesIO()
                try:
                    with Image.open(BytesIO(raw)) as source:
                        source.verify()
                    with Image.open(BytesIO(raw)) as source:
                        oriented = ImageOps.exif_transpose(source)
                        try:
                            with oriented.convert("RGB") as rgb_source:
                                fitted = ImageOps.fit(
                                    rgb_source,
                                    (72, 72),
                                    method=Image.Resampling.LANCZOS,
                                    centering=(0.5, 0.5),
                                )
                            try:
                                fitted.save(converted, "PNG")
                            finally:
                                fitted.close()
                        finally:
                            if oriented is not source:
                                oriented.close()
                    converted.seek(0)
                    picture = ExcelImage(converted)
                except (UnidentifiedImageError, OSError) as exc:
                    converted.close()
                    raise AppError("الصورة غير صالحة.") from exc
                keepalive.append(converted)
                picture.anchor = OneCellAnchor(
                    _from=AnchorMarker(col=0, colOff=0, row=row - 1, rowOff=0),
                    ext=XDRPositiveSize2D(cx=pixels_to_EMU(72), cy=pixels_to_EMU(72)),
                )
                sheet.add_image(picture)
                name_cell = sheet.cell(row, 2, item.product_name)
                name_cell.font = Font(name="Arial", size=14, bold=True, color="000000")
                quantity_cell = sheet.cell(row, 3, item.quantity)
                quantity_cell.font = Font(name="Arial", size=13, bold=True, color="000000")
                sheet.row_dimensions[row].height = 54
                for col in range(1, 4):
                    sheet.cell(row, col).alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                        readingOrder=2 if col == 2 else 0,
                    )
                    sheet.cell(row, col).border = border
            sheet.column_dimensions["A"].width = 9.5
            sheet.column_dimensions["B"].width = 25
            sheet.column_dimensions["C"].width = 9
            sheet.freeze_panes = "A4"
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        finally:
            workbook.close()
            for buffer in keepalive:
                buffer.close()


def safe_excel_filename(title):
    clean = re.sub(r'[\\/:*?"<>|\x00-\x1f\x7f]', "-", title).strip(" .-")[:100]
    return f"{clean or 'order'}.xlsx"
