import asyncio
import copy
import logging
import math
import re
import zipfile
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from pathlib import PurePosixPath

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.errors import ValidationError

logger = logging.getLogger(__name__)

OFFICE_RATE = Decimal("0.03")
RESULT_HEADERS = ("السعر بالأوقية", "نسبة المكتب", "الشحن", "سعر الطياح")
ALIASES = {
    "PRICE": {"PRICE", "UNIT PRICE", "RMB PRICE", "PRICE RMB", "UNIT PRICE RMB"},
    "PCS": {"PCS", "PCS CTN", "PCS PER CTN", "QTY CTN", "PIECES CTN"},
    "CBM": {"CBM", "CBM CTN", "CARTON CBM", "CBM PER CTN"},
}


def normalize_header(value):
    text = str(value if value is not None else "").replace("\xa0", " ")
    text = re.sub(r"[\r\n]+", " ", text.strip().upper())
    text = re.sub(r"[^\w\u0600-\u06ff]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


class ExcelPricingService:
    def __init__(self, settings):
        self.settings = settings

    async def transform(self, source_bytes, rmb_rate, shipping_cost_per_cbm, filename=None):
        rate = self._parameter(rmb_rate, "معامل تحويل RMB غير صالح.")
        shipping_cost = self._parameter(shipping_cost_per_cbm, "تكلفة الشحن لكل CBM غير صالحة.")
        self._validate_archive(source_bytes)
        return await asyncio.to_thread(
            self._transform_sync, source_bytes, rate, shipping_cost, filename
        )

    @staticmethod
    def _parameter(raw, message):
        text = str(raw if raw is not None else "").strip()
        if not text or len(text) > 64:
            raise ValidationError(message)
        try:
            value = Decimal(text)
        except InvalidOperation as exc:
            raise ValidationError(message) from exc
        if not value.is_finite() or value <= 0 or len(value.as_tuple().digits) > 30:
            raise ValidationError(message)
        return value

    def _validate_archive(self, data):
        if len(data) > self.settings.max_excel_upload_mb * 1024 * 1024:
            raise ValidationError("حجم ملف Excel يتجاوز الحد المسموح.")
        try:
            with zipfile.ZipFile(BytesIO(data)) as archive:
                infos = archive.infolist()
                names = {item.filename for item in infos}
                if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                if len(infos) > self.settings.max_zip_entries:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                limit = self.settings.max_uncompressed_import_mb * 1024 * 1024
                if sum(item.file_size for item in infos) > limit:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                for item in infos:
                    path = PurePosixPath(item.filename)
                    if item.flag_bits & 1 or path.is_absolute() or ".." in path.parts:
                        raise ValidationError("الملف تالف أو غير مدعوم.")
                    if "\\" in item.filename:
                        raise ValidationError("الملف تالف أو غير مدعوم.")
                bad = archive.testzip()
                if bad:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
        except (zipfile.BadZipFile, OSError) as exc:
            raise ValidationError("الملف تالف أو غير مدعوم.") from exc

    def _transform_sync(self, data, rate, shipping_cost, filename=None):
        formula_book = value_book = None
        sheet_name = None
        try:
            formula_book = load_workbook(BytesIO(data), data_only=False, keep_links=True)
            value_book = load_workbook(BytesIO(data), data_only=True, keep_links=True)
            processed = 0
            for sheet in formula_book.worksheets:
                sheet_name = sheet.title
                value_sheet = value_book[sheet.title]
                header = self._find_header(sheet)
                if header is None:
                    continue
                row_number, sources = header
                result_columns = self._result_columns(sheet, row_number)
                self._calculate_rows(
                    sheet, value_sheet, row_number, sources, result_columns, rate, shipping_cost
                )
                processed += 1
            if not processed:
                raise ValidationError("لم يتم العثور على أعمدة PRICE وPCS وCBM في الملف.")
            output = BytesIO()
            formula_book.save(output)
            return output.getvalue()
        except ValidationError:
            raise
        except Exception as exc:
            logger.exception(
                "Unexpected Excel pricing failure (filename=%r, sheet=%r)",
                filename,
                sheet_name,
            )
            raise ValidationError("تعذر معالجة ملف Excel بسبب خطأ غير متوقع.") from exc
        finally:
            if formula_book is not None:
                formula_book.close()
            if value_book is not None:
                value_book.close()

    def _find_header(self, sheet):
        complete = []
        partial = []
        for row in range(1, min(50, sheet.max_row) + 1):
            matches = {key: [] for key in ALIASES}
            for cell in sheet[row]:
                normalized = normalize_header(cell.value)
                for key, aliases in ALIASES.items():
                    if normalized in aliases:
                        matches[key].append(cell.column)
            present = [key for key, columns in matches.items() if columns]
            if present:
                partial.append((row, matches))
            if len(present) == 3:
                for key, columns in matches.items():
                    if len(columns) != 1:
                        raise ValidationError(
                            f"تعذر تحديد عمود {key} بشكل فريد في الورقة {sheet.title}."
                        )
                columns = {key: values[0] for key, values in matches.items()}
                if len(set(columns.values())) != 3:
                    raise ValidationError(f"تعذر تحديد أعمدة البيانات في الورقة {sheet.title}.")
                complete.append((row, columns))
        if len(complete) > 1:
            raise ValidationError(f"تم العثور على أكثر من صف عناوين في الورقة {sheet.title}.")
        if complete:
            return complete[0]
        if partial:
            best_row, matches = max(
                partial, key=lambda item: sum(bool(x) for x in item[1].values())
            )
            del best_row
            missing = [key for key, columns in matches.items() if not columns]
            if len(missing) == 1:
                raise ValidationError(f"الورقة {sheet.title}: العمود {missing[0]} مفقود.")
        return None

    def _result_columns(self, sheet, row):
        normalized_results = {normalize_header(value): value for value in RESULT_HEADERS}
        found = {value: [] for value in RESULT_HEADERS}
        last_real = 0
        for cell in sheet[row]:
            if cell.value is not None and str(cell.value).strip():
                last_real = cell.column
                result = normalized_results.get(normalize_header(cell.value))
                if result:
                    found[result].append(cell.column)
        for result, columns in found.items():
            if len(columns) > 1:
                raise ValidationError(f"عنوان النتيجة {result} مكرر في الورقة {sheet.title}.")
        existing = [found[result][0] for result in RESULT_HEADERS if found[result]]
        if existing and len(existing) != len(RESULT_HEADERS):
            raise ValidationError(
                f"عناوين النتائج غير مكتملة في الورقة {sheet.title}؛ لن تتم الكتابة فوق البيانات."
            )
        if existing:
            if not self._columns_safe(sheet, existing, allow_result_headers=True, header_row=row):
                raise ValidationError(f"أعمدة النتائج الموجودة غير آمنة في الورقة {sheet.title}.")
            return dict(zip(RESULT_HEADERS, existing, strict=True))

        # max_column includes cells which merely have formatting. Inspect actual values instead.
        last_real = max(
            (
                cell.column
                for sheet_row in sheet.iter_rows()
                for cell in sheet_row
                if self._has_value(cell.value)
            ),
            default=0,
        )
        start = last_real + 1
        while not self._columns_safe(sheet, range(start, start + 4)):
            start += 1
        result_columns = {}
        template = sheet.cell(row, last_real) if last_real else None
        for offset, result in enumerate(RESULT_HEADERS):
            column = start + offset
            cell = sheet.cell(row, column, result)
            if template:
                self._copy_style(template, cell)
            sheet.column_dimensions[get_column_letter(column)].width = 18
            result_columns[result] = column
        return result_columns

    @staticmethod
    def _has_value(value):
        return value is not None and (not isinstance(value, str) or bool(value.strip()))

    def _columns_safe(self, sheet, columns, allow_result_headers=False, header_row=None):
        columns = set(columns)
        for merged in sheet.merged_cells.ranges:
            if columns.intersection(range(merged.min_col, merged.max_col + 1)):
                return False
        for image in sheet._images:
            anchor = image.anchor
            if hasattr(anchor, "_from"):
                end = getattr(getattr(anchor, "to", None), "col", anchor._from.col)
                image_columns = range(anchor._from.col + 1, end + 2)
                if columns.intersection(image_columns):
                    return False
        allowed = {normalize_header(value) for value in RESULT_HEADERS}
        for column in columns:
            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row, column)
                header_allowed = (
                    allow_result_headers
                    and row == header_row
                    and normalize_header(cell.value) in allowed
                )
                value_forbidden = self._has_value(cell.value) and not (
                    header_allowed or allow_result_headers
                )
                if value_forbidden or cell.comment or cell.hyperlink:
                    return False
        return True

    def _calculate_rows(self, sheet, values, header_row, sources, results, rate, shipping_cost):
        for row in range(header_row + 1, sheet.max_row + 1):
            raw = {key: values.cell(row, column).value for key, column in sources.items()}
            formula = {key: sheet.cell(row, column).value for key, column in sources.items()}
            if not any(value is not None and str(value).strip() for value in formula.values()):
                continue
            row_text = " ".join(
                normalize_header(sheet.cell(row, column).value)
                for column in range(1, min(sheet.max_column, 12) + 1)
            )
            if re.search(r"\b(TOTAL|GRAND TOTAL|SUBTOTAL)\b", row_text):
                continue
            parsed = {}
            for key in ("PRICE", "PCS", "CBM"):
                value = raw[key]
                parsed[key] = self._cell_decimal(value, key, sheet.title, row)
            price, pcs, cbm = parsed["PRICE"], parsed["PCS"], parsed["CBM"]
            price_ouguiya = price * rate if price is not None else None
            office_fee = price_ouguiya * OFFICE_RATE if price_ouguiya is not None else None
            shipping = cbm * shipping_cost / pcs if cbm is not None and pcs and pcs > 0 else None
            available = [
                value for value in (price_ouguiya, office_fee, shipping) if value is not None
            ]
            calculated = (
                price_ouguiya,
                office_fee,
                shipping,
                sum(available, Decimal(0)) if available else None,
            )
            adjacent = sheet.cell(row, max(sources.values()))
            for heading, value in zip(RESULT_HEADERS, calculated, strict=True):
                target = sheet.cell(row, results[heading])
                if not target.has_style:
                    self._copy_style(adjacent, target)
                if value is None:
                    target.value = None
                else:
                    rounded = value.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
                    target.value = float(rounded)
                    target.number_format = "0.######"

    @staticmethod
    def _cell_decimal(raw, name, sheet, row):
        if raw is None:
            return None
        if isinstance(raw, str) and raw.strip().upper() in {"", "-", "--", "N/A", "NA"}:
            return None
        try:
            if isinstance(raw, bool):
                raise InvalidOperation
            if isinstance(raw, float) and not math.isfinite(raw):
                raise InvalidOperation
            value = Decimal(str(raw).strip().replace(",", ""))
        except (InvalidOperation, ValueError):
            logger.warning(
                "Non-numeric pricing value (sheet=%r, row=%d, column=%s)", sheet, row, name
            )
            return None
        if not value.is_finite() or value < 0:
            logger.warning("Invalid pricing value (sheet=%r, row=%d, column=%s)", sheet, row, name)
            return None
        return value

    @staticmethod
    def _copy_style(source, target):
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        target.number_format = source.number_format
